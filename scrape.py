#!/usr/bin/env python3
"""
Scrape Estatutos y Reglamentos de Pregrado from UDEA Normativa.
- Downloads all PDFs linked from the table filtered by
  "REGLAMENTO ESTUDIANTIL DE PREGRADO" into ./pdfs/
- Saves a normalized Excel at ./estatutos_reglamentos_pregrado.xlsx
"""
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "https://normativa.udea.edu.co"
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
})

ROOT = Path("/Users/laalquimia/Projects/udeaScrape")
PDF_DIR = ROOT / "pdfs"
PDF_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = ROOT / "estatutos_reglamentos_pregrado.xlsx"
JSON_PATH = ROOT / "data.json"

# Data extracted from the live normativa.udea.edu.co table
# filtered by "REGLAMENTO ESTUDIANTIL DE PREGRADO" (28 records, 1 page)
RECORDS = [
    {"numero": "623",  "docId": "35137054",  "fechaExpedicion": "2025/03/28", "entradaVigencia": "2025/03/28", "medioPublicacion": "NORMATIVA.UDEA.EDU.CO", "resuelve": "REGLAMENTAR LA RUTA CURRICULAR ESPECIAL DE DOBLE TITULACIÓN ENTRE LOS PROGRAMAS DE LICENCIATURA EN EDUCACIÓN FÍSICA Y ENTRENAMIENTO DEPORTIVO, ADSCRITOS AL INSTITUTO UNIVERSITARIO DE EDUCACIÓN FÍSICA Y DEPORTE.", "normasRelacionadas": ""},
    {"numero": "0425", "docId": "17917413",  "fechaExpedicion": "2014/07/29", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR LOS ARTICULOS 130 Y 216 DEL ACUERDO SUPERIOR 01 DE 1981 REGLAMENTO ESTUDIANTIL DE PREGRADO NUMERO MINIMO DE CREDITOS PARA ESTIMULOS ACADEMICOS", "normasRelacionadas": "ACUERDO SUPERIOR 01 DE 1981"},
    {"numero": "0385", "docId": "7753259",   "fechaExpedicion": "2010/11/30", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR ARTICULOS 49 50 51 53 Y 180 DEL ACUERDO SUPERIOR 01 DEL 15 DE FEBRERO DE 1981 REGLAMENTO ESTUDIANTIL DE PREGRADO REAJUSTE CONDICION DOBLE TITULACION", "normasRelacionadas": "ACUERDO SUPERIOR 01 DEL 15 DE FEBRERO DE 1981"},
    {"numero": "0376", "docId": "6702815",   "fechaExpedicion": "2009/11/24", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICA ARTICULO 63 DEL ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981 REGLAMENTO ESTUDIANTIL PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981"},
    {"numero": "0239", "docId": "491602",    "fechaExpedicion": "2002/12/03", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR ARTICULO 216 DEL  ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL  DE PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0229", "docId": "500235",    "fechaExpedicion": "2002/07/02", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR EL ARTICULO 163 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0180", "docId": "122606",    "fechaExpedicion": "2000/10/12", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR ARTICULO 215 DEL  ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981 REGLAMENTO ESTUDIANTIL Y NORMAS ACADEMICAS", "normasRelacionadas": "ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981"},
    {"numero": "0181", "docId": "122609",    "fechaExpedicion": "2000/10/12", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR EL ARTICULO 73 DEL ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981 REGLAMENTO ESTUDIANTIL PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DEL 15 DE FEBRERO DE 1981"},
    {"numero": "0170", "docId": "122524",    "fechaExpedicion": "2000/02/03", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR LOS ARTICULOS 77 Y 78 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL DE PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0164", "docId": "122483",    "fechaExpedicion": "1999/12/16", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR ACUERDO SUPERIOR 1 DE 1981 Y DEROGAR ACUERDO SUPERIOR 177 DE 1991 REGLAMENTO ESTUDIANTIL PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981 Y ACUERDO SUPERIOR 177 DE 1991"},
    {"numero": "0165", "docId": "122487",    "fechaExpedicion": "1999/12/16", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "SUPRIMIR EL PARAGRAFO 2 DEL ARTICULO 82 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0146", "docId": "123140",    "fechaExpedicion": "1998/08/10", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR UN PARAGRAFO AL ARTICULO 170 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0137", "docId": "123017",    "fechaExpedicion": "1998/05/11", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "SUPRIMIR PARTE FINAL ARTICULO 124 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL DE PREGRADO Y DE NORMAS ACADEMICAS", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0118", "docId": "122662",    "fechaExpedicion": "1997/07/07", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR UN PARAGRAFO AL ARTICULO 74 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS SOBRE CANCELACION DE CURSOS", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0038", "docId": "121787",    "fechaExpedicion": "1995/03/13", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR EL ARTICULO 172 DEL REGLAMENTO ESTUDIANTIL DE PREGRADO", "normasRelacionadas": ""},
    {"numero": "0034", "docId": "121776",    "fechaExpedicion": "1995/02/27", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "SUSTITUIR EL ARTICUOLO 171 DEL REGLAMENTO ESTUDIANTIL DE PREGRADO", "normasRelacionadas": ""},
    {"numero": "0033", "docId": "121772",    "fechaExpedicion": "1995/02/13", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR LOS ARTICULOS 221 222 Y 223 DEL ACUERDO 1 DE 1981 REGLAMENTO ESTUDIANTIL DE PREGRADO PREMIO INVESTIGACION", "normasRelacionadas": "ACUERDO 1 DE 1981"},
    {"numero": "0031", "docId": "121775",    "fechaExpedicion": "1995/01/25", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR REGLAMENTO ESTUDIANTIL Y NORMAS ACADEMCIAS DE PREGRADO", "normasRelacionadas": ""},
    {"numero": "0267", "docId": "323317",    "fechaExpedicion": "1993/09/06", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR EL ARTICULO 72 DEL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL Y NORMAS ACADEMICAS", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0225", "docId": "330501",    "fechaExpedicion": "1992/08/18", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "MODIFICAR EL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0177", "docId": "328304",    "fechaExpedicion": "1991/04/15", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR EL ACUERDO SUPERIOR 1 DE 1981 REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS SOBRE DERECHO DE REINGRESO PROGRAMAS DE PREGRADO", "normasRelacionadas": "ACUERDO SUPERIOR 1 DE 1981"},
    {"numero": "0010", "docId": "261408",    "fechaExpedicion": "1984/06/06", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "EXCEPCIONAR TEMPORALMENTE ALGUNOS ARTICULOS REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICA", "normasRelacionadas": ""},
    {"numero": "0025", "docId": "257115",    "fechaExpedicion": "1983/06/01", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "ADICIONAR ACUERDO 1 DE FEBRERO 15 DE 1981 CONSEJO SUPERIOR REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS SOBRE CERTIFICACION DE CONOCIMIENTOS", "normasRelacionadas": "ACUERDO 1 DE FEBRERO 15 DE 1981 CONSEJO SUPERIOR"},
    {"numero": "0027", "docId": "257207",    "fechaExpedicion": "1983/06/01", "entradaVigencia": "",            "medioPublicacion": "",                            "resuelve": "REGLAMENTAR EL ARTICULO 25 DEL REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS TITULO POST MORTEM", "normasRelacionadas": ""},
    {"numero": "1",    "docId": "35187147",  "fechaExpedicion": "1981/02/15", "entradaVigencia": "1981/02/15", "medioPublicacion": "NORMATIVA.UDEA.EDU.CO", "resuelve": "EXPIDE EL REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS VERSION ACTUALIZADA POR LA SECRETARIA GENERAL A 15 DE DICIEMBRE DE 2025 SIN CONCORDANCIAS", "normasRelacionadas": ""},
    {"numero": "1",    "docId": "31878551",  "fechaExpedicion": "1981/02/15", "entradaVigencia": "1981/02/15", "medioPublicacion": "NORMATIVA.UDEA.EDU.CO", "resuelve": "EXPIDE EL REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS", "normasRelacionadas": ""},
    {"numero": "1",    "docId": "35187099",  "fechaExpedicion": "1981/02/15", "entradaVigencia": "1981/02/15", "medioPublicacion": "NORMATIVA.UDEA.EDU.CO", "resuelve": "EXPIDE EL REGLAMENTO ESTUDIANTIL Y DE NORMAS ACADEMICAS VERSION ACTUALIZADA POR LA SECRETARIA GENERAL A 15 DE DICIEMBRE DE 2025 CON CONCORDANCIAS", "normasRelacionadas": ""},
    {"numero": "0003", "docId": "415356",    "fechaExpedicion": "1980/11/18", "entradaVigencia": "",            "medioPublicacion": "NORMATIVA.UDEA.EDU.CO", "resuelve": "CONCEPTUAR FAVORABLEMENTE LA ADOPCION DEL REGLAMENTO ESTUDIANTIL NORMAS ACADEMICAS Y DISCIPLINARIAS", "normasRelacionadas": ""},
]


def sanitize_filename(name: str) -> str:
    name = name.strip() or "sin_numero"
    name = re.sub(r"[^\w\-. ]+", "_", name, flags=re.UNICODE)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:120] if name else "doc"


def fetch_extension(doc_id: str) -> tuple[str, str]:
    """Return (extension, codigoimagen) from ExtensionDocumento endpoint."""
    r = SESSION.post(
        f"{BASE}/Documentos/ExtensionDocumento",
        data={"codigodocumento": doc_id},
        headers={"X-Requested-With": "XMLHttpRequest"},
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    return data.get("extension", "pdf"), str(data.get("codigoimagen", ""))


def download_pdf(rec: dict, retries: int = 3) -> dict:
    doc_id = rec["docId"]
    numero = rec["numero"]
    fecha = rec["fechaExpedicion"].replace("/", "-")
    # Friendly filename: Numero_FechaExpedicion_<docId>.pdf
    base = f"Acuerdo_{numero}_{fecha}_{doc_id}"
    filename = sanitize_filename(base) + ".pdf"
    out = PDF_DIR / filename

    for attempt in range(1, retries + 1):
        try:
            ext, codigoimagen = fetch_extension(doc_id)
            if ext.lower() != "pdf":
                raise RuntimeError(f"Documento {doc_id} no es PDF (extension={ext})")
            url = f"{BASE}/Documentos/Documento?codigodocumento={doc_id}&codigoimagen={codigoimagen}&buscarpdf="
            r = SESSION.get(url, timeout=120, allow_redirects=True)
            r.raise_for_status()
            if not r.content.startswith(b"%PDF"):
                raise RuntimeError(f"Respuesta no es PDF para {doc_id} (len={len(r.content)})")
            out.write_bytes(r.content)
            rec_out = dict(rec)
            rec_out.update({
                "archivo": out.name,
                "tamano_bytes": out.stat().st_size,
                "url_pdf": url,
            })
            return rec_out
        except Exception as e:
            print(f"  [intento {attempt}/{retries}] {numero} ({doc_id}): {e}", file=sys.stderr)
            time.sleep(1.5 * attempt)
    rec_out = dict(rec)
    rec_out.update({"archivo": "", "tamano_bytes": 0, "url_pdf": "", "error": "fallo_descarga"})
    return rec_out


def build_excel(records: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estatutos Reglamentos Pregrado"

    headers = [
        "Número", "Fecha de expedición", "Entrada en vigencia",
        "Medio de publicación", "Resuelve", "Normas relacionadas",
        "Archivo PDF", "Tamaño (bytes)", "URL PDF", "ID documento",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E2C")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in records:
        ws.append([
            r.get("numero", ""),
            r.get("fechaExpedicion", ""),
            r.get("entradaVigencia", ""),
            r.get("medioPublicacion", ""),
            r.get("resuelve", ""),
            r.get("normasRelacionadas", ""),
            r.get("archivo", ""),
            r.get("tamano_bytes", 0),
            r.get("url_pdf", ""),
            r.get("docId", ""),
        ])

    # Column widths
    widths = [10, 14, 14, 22, 60, 35, 35, 14, 50, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap text on data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    # Freeze header
    ws.freeze_panes = "A2"
    # Add auto filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main() -> int:
    print(f"Iniciando descarga de {len(RECORDS)} PDFs en {PDF_DIR} ...")
    enriched = []
    for i, rec in enumerate(RECORDS, 1):
        numero = rec["numero"]
        print(f"[{i:02d}/{len(RECORDS)}] Acuerdo {numero} (id={rec['docId']}) ...", end=" ", flush=True)
        result = download_pdf(rec)
        if result.get("archivo"):
            print(f"OK -> {result['archivo']} ({result['tamano_bytes']} bytes)")
        else:
            print("FAIL")
        enriched.append(result)

    JSON_PATH.write_text(json.dumps(enriched, indent=2, ensure_ascii=False))
    print(f"\nDatos guardados en {JSON_PATH}")

    build_excel(enriched, EXCEL_PATH)
    print(f"Excel generado en {EXCEL_PATH}")

    ok = sum(1 for r in enriched if r.get("archivo"))
    print(f"\nResumen: {ok}/{len(enriched)} PDFs descargados correctamente.")
    print("Archivos en carpeta pdfs/:")
    for p in sorted(PDF_DIR.iterdir()):
        print(f"  - {p.name}  ({p.stat().st_size:,} bytes)")
    return 0 if ok == len(enriched) else 1


if __name__ == "__main__":
    sys.exit(main())
